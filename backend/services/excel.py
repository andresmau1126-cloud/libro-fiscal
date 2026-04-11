"""
Servicio de exportación a Excel.
RF-008: Exportación con formato profesional.
"""
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

MES_ES = [
    None, "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]


def generate_excel(libro, rows, month=None):
    """
    Genera un archivo Excel (.xlsx) en memoria con los movimientos.
    Retorna BytesIO listo para enviar.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Fiscal"

    bold = Font(b=True)
    title_font = Font(b=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "LIBRO FISCAL DE REGISTRO DE OPERACIONES DIARIAS"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    # Header
    mes_nombre = MES_ES[int(month)] if month else ""
    ws["A3"] = "Nombre:"
    ws["A3"].font = bold
    ws["A3"].alignment = left
    ws["B3"] = libro.get("nombre", "")
    ws.merge_cells("B3:C3")
    ws["D3"] = "NIT:"
    ws["D3"].font = bold
    ws["D3"].alignment = right
    ws["E3"] = libro.get("nit", "")

    ws["A4"] = "Mes:"
    ws["A4"].font = bold
    ws["A4"].alignment = left
    ws["B4"] = mes_nombre
    ws.merge_cells("B4:C4")
    ws["D4"] = "Año:"
    ws["D4"].font = bold
    ws["D4"].alignment = right
    ws["E4"] = libro.get("anio", "")

    ws.append([])

    # Table headers
    headers = ["DIA", "DESCRIPCIÓN", "INGRESOS", "EGRESOS", "SALDO"]
    hdr_row = ws.max_row + 1
    ws.append(headers)
    for c in ws[hdr_row]:
        c.font = bold
        c.alignment = center
        c.border = border

    # Data rows
    for r in rows:
        fecha = r.get("fecha")
        if isinstance(fecha, (datetime, date)):
            dia = fecha.day
        else:
            dia = int(str(fecha)[8:10])
        ws.append([
            dia,
            r["descripcion"],
            float(r["ingresos"] or 0),
            float(r["egresos"] or 0),
            float(r["saldo"] or 0),
        ])

    data_end = ws.max_row

    # Totals
    total_ing = sum(float(r["ingresos"] or 0) for r in rows)
    total_egr = sum(float(r["egresos"] or 0) for r in rows)
    total_sal = float(rows[-1]["saldo"]) if rows else 0.0

    total_row = data_end + 1
    ws.cell(row=total_row, column=2, value="TOTAL").font = bold
    ws.cell(row=total_row, column=3, value=total_ing).font = bold
    ws.cell(row=total_row, column=4, value=total_egr).font = bold
    ws.cell(row=total_row, column=5, value=total_sal).font = bold

    # Borders and number format
    for row in ws.iter_rows(min_row=hdr_row, max_row=total_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = border
            if cell.column in (3, 4, 5) and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
            if cell.column == 1:
                cell.alignment = center
            if cell.column in (3, 4, 5):
                cell.alignment = right

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
